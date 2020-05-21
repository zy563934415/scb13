#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2020/5/20 19:25
# @Author : ye

'''
1、用例，读取测试数据
2、用数据发送接口请求，得到执行结果
3、用执行结果 vs 实际结果 --对比，得出结论
4、得出结果，回写到测试用例

'''

import openpyxl
import requests
session = requests.session() # 自动带入token，不用手动输入了
# 函数  --读取测试用例数据的
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename) # 加载工作簿, --excel表格 ---赋值给wb这个变量
    sheet = wb[sheetname] # 表单 --通过工作簿[]取表单
    max_row = sheet.max_row # 获取最大行号
    cases = []  # 空列表 --后续传值
    for i in range(2,max_row+1):  # 取头不取尾，所以需要+1
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,  # 获取case_id
        url = sheet.cell(row=i,column=5).value,      # 获取url
        data = sheet.cell(row=i,column=6).value,     # 获取数据
        expected_result = sheet.cell(row=i,column=7).value # 获取期望结果
        )   # 一个用例存放到一个字典
        cases.append(case)  # 将字典中的数据，追加到列表里面存储
    # print(cases)
    return cases

# 函数  --用来发送接口请求的
def post_func(qcd_url,qcd_data):    # 地址和数据
    # res = requests.post(qcd_url,qcd_data)    # post 方式发送接口请求
    res = session.post(qcd_url,qcd_data)    # session来发送接口请求，会自动带上cookie或token
    result = res.json()  # 字典
    return result   # 返回值  --响应消息的结果

# 函数  --写入结果
# filename -excel文件名, sheetname -表单名, row -行,column -列,real_result -实际结果
def write_result(filename,sheetname,row,column,real_result):
    wb = openpyxl.load_workbook(filename) # excel文件里面
    sheet = wb[sheetname]   # 表单
    sheet.cell(row=row,column=column).value = real_result # 修改某一行的某一列的单元格中的数据
    wb.save(filename) # 保存

# 执行
def execute_func(filename,sheetname): #filename -excel文件名, sheetname -表单名
    test_cases = read_data(filename,sheetname) # 调用read_data函数 --读取测试用例数据的
    for case in test_cases: # 进行遍历,以字典方式存储
        case_id = case.get('case_id')   # 获取到对应的case_id
        url = case["url"]   # 获取到对应的url
        data = case["data"] # 获取到对应的参数 --由于data从excel表格里面得到的是str字符串格式的
        data = eval(data)   # 所以需要转换为字典格式，eval()进行转换 字符串-->字典
        expected_result = case.get("expected_result")   # 获取到对应的预期结果
        expected_result = expected_result.replace("null","None")    # 替换 -将字符串中的 null 替换为 None
        expected_result = eval(expected_result) # 转换 字符串-->字典 在进行转换是发现null,会提示错误识别不到null
        real_response = post_func(url, data)    # 调用 -发送接口请求
        expected_msg = expected_result.get("msg") # 获取到预期结果
        real_msg = real_response.get("msg") # 获取到实际结果

        print('真实执行结果是：{}'.format(real_msg))
        print('预期测试结果是：{}'.format(expected_msg))
        if real_msg == expected_msg: # 预期结果 对比 实际结果
            print("第{}条测试用例通过".format(case_id))
            final_result = "Passed" # 定义变量  --目的：返写结果
        else:
            print("第{}条测试用例不通过".format(case_id))
            final_result = "Failed"
        print("**" * 20)
        write_result(filename,sheetname,case_id+1,8,final_result) # 调用了回写的函数  --进行结果的写入

execute_func("test_case.xlsx","recharge") # 调用执行函数